import time
import pandas as pd
import openpyxl
from dotenv import load_dotenv
import datetime
import pygetwindow as gw
import pyautogui
import os
from send_email import send_email
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from baixar_fatura import extrair_dados
import time
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook 

load_dotenv()

data_atual = datetime.date.today()
mes_ano = data_atual.strftime("%m-%Y")
hora_minuto = datetime.datetime.now().strftime('%H:%M')
download_directory = os.getenv("DOWNLOAD_DIRECTORY")

if __name__ == "__main__":
    subject = 'RPA - Conta de Consumo'    
    body = 'Processo iniciado'
    attachment_path = ''
    send_email(subject, body, attachment_path)


pasta_destino = os.getenv("PASTA_DESTINO")
planilha_original = pd.read_excel(os.getenv("PLANILHA_ORIGINAL"))
relatorio_atual = planilha_original.copy()
nome_relatorio = "Relatório de execução " + mes_ano +".xlsx"
caminho_completo = os.path.join(pasta_destino, nome_relatorio)
relatorio_atual.to_excel(caminho_completo, index=False)
print(nome_relatorio)
print(caminho_completo)

workbook = openpyxl.load_workbook(os.getenv('PATH_EXCEL_LOGIN'))
sheet_produtos = workbook['login']
quantidade_de_linhas = sheet_produtos.max_row

#Define resolução da tela

largura, altura = 1920, 1080
janela_principal = gw.getWindowsWithTitle('')[0]
janela_principal.size = (largura, altura)

 # Configurações do Chrome
chrome_options = ChromeOptions()
#chrome_options.add_argument('--headless')
chrome_options.add_argument("--start-maximized")
chrome_options.add_argument("--disable-popup-blocking")
chrome_options.add_argument("--disable-notifications")
chrome_options.add_argument("--disable-infobars")
chrome_options.add_experimental_option('prefs', {
    'download.default_directory': download_directory,
    'download.prompt_for_download': False,
    'download.directory_upgrade': True,
    'safebrowsing.enabled': True
})
# Abrir o Chrome
url = os.getenv("URL")
browser = webdriver.Chrome(options=chrome_options)
#url = url  # Substitua pelo URL desejado
browser.get(url)
wait = WebDriverWait(browser, 30)

WebDriverWait(browser, 60).until(EC.text_to_be_present_in_element((By.TAG_NAME, 'body'), 'Que bom te ver por aqui!'))

time.sleep(2)
pyautogui.click(960,596,duration=0.2)


for linha in sheet_produtos.iter_rows(min_row=2, values_only=True):
    
    if not linha[0]:
        continue
    
    wait = WebDriverWait(browser, 60)
    login = wait.until(EC.visibility_of_element_located((By.ID, 'email'))).send_keys(linha[0])
    password = wait.until(EC.visibility_of_element_located((By.ID, 'senha'))).send_keys(linha[1])
    button_entrar = wait.until(EC.presence_of_element_located((By.XPATH, f'//*[@translate="@APP-LOGIN-ENTRAR"]'))).click()
    
    button_entrar = wait.until(EC.presence_of_element_located((By.XPATH, f'//*[@translate="@APP-COMMON-ENTRAR"]'))).click()
    #Continuar daqui para baixo, implementar logica as filiais

    valida_referencia = wait.until(EC.presence_of_element_located((By.XPATH, '//*[contains(@translate, "@APP-PORTAL-CARD-CONTA-ATUAL-VENCIMENTO-REFERENTE")]'))).text
    print(valida_referencia)

    valida_status = wait.until(EC.presence_of_element_located((By.XPATH, f'//*[contains(@ng-bind, "$ctrl.portal.contaAtual.situacao")]'))).text
    print(valida_status)

    button_download = wait.until(EC.presence_of_element_located((By.XPATH, f'//*[contains(@translate, "@APP-PORTAL-MODAL-BAIXAR-CONTA")]'))).click()

    class MyHandler(FileSystemEventHandler):
        def __init__(self, file_path):
            self.file_path = file_path
            self.file_created = False

        def on_created(self, event):
            if event.src_path == self.file_path:
                self.file_created = True

    # Substitua o caminho do arquivo conforme necessário
    caminho_do_arquivo = os.getenv('PASTA_CRIACAO_ARQUIVO')

    # Cria um manipulador de eventos
    event_handler = MyHandler(file_path=caminho_do_arquivo)

    # Cria um observador para monitorar o diretório
    observer = Observer()
    observer.schedule(event_handler, path=caminho_do_arquivo, recursive=False)
    observer.start()

    # Aguarda por 5 segundos ou até que o arquivo seja criado
    tempo_maximo_espera = 10
    inicio_tempo = time.time()

    while time.time() - inicio_tempo < tempo_maximo_espera and not event_handler.file_created:
        time.sleep(1)

    # Para o observador
    observer.stop()
    observer.join()

    # Verifica se o arquivo foi criado
    if event_handler.file_created:
        print("O arquivo foi criado.")
    else:
        print("O tempo máximo de espera foi atingido e o arquivo não foi criado.")
    
    #time.sleep(10)

    if __name__ == "__main__":
        #valores = ['Enel', instalacao, referencia, emissao, vencimento, valor]
        extrair_dados() 
    
    button_sair = wait.until(EC.presence_of_element_located((By.XPATH, f'//*[contains(@translate, "@APP-COMMON-SAIR")]'))).click()
    #browser.quit()

    hora_minuto = datetime.datetime.now().strftime('%H:%M')
    #nome_relatorio = "Relatório de execução " + mes_ano +".xlsx"

    path_dados = caminho_completo
    workbook = load_workbook(path_dados)
    sheet = workbook['Sheet1']
    valores = [data_atual, hora_minuto, 'Enel', 'Captura fatura', 'Baixada']
    sheet.append(valores)
    workbook.save(caminho_completo)

if __name__ == "__main__":
    subject = 'RPA - Conta de Consumo'    
    body = 'Processo Finalizado'
    attachment_path = caminho_completo
    send_email(subject, body,attachment_path)
